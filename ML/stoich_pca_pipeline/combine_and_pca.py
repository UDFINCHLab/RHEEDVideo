"""
Script 3: combine all preprocessed H5 files into one master H5,
attach labels from the manifest, run PCA, find optimal K, then run K-Means.
Edit DATASET_NAME and TARGET_COLUMN below to switch datasets.
"""

import sys
import re
from pathlib import Path

import h5py as h5
import numpy as np
import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
ML_DIR = PROJECT_ROOT / "ML"
sys.path.insert(0, str(ML_DIR))

from pca import pca  # noqa: E402
from k_means import k_means  # noqa: E402


# -----------------------------
# DATASET SELECTION — edit these two lines to switch datasets
# -----------------------------
DATASET_NAME  = "CSO"           # "STO" or "CSO"
TARGET_COLUMN = "lattice parameter (out-of-plane)"  # "XPS value" for STO, lattice param for CSO


# -----------------------------
# PATHS
# -----------------------------
# PREPROCESSED_DIR: folder containing the individual H5 files from batch_preprocess.py
# MANIFEST_FILE:    Excel or CSV file mapping sample IDs to target values
# COMBINED_H5:      output master H5 combining all videos — saved under results/DATASET_NAME/
PREPROCESSED_DIR = PROJECT_ROOT / "stoich_dataset" / DATASET_NAME / "preprocessed_h5"
MANIFEST_FILE    = PROJECT_ROOT / "stoich_dataset" / DATASET_NAME / "labels" / "CSO log.xlsx"
OUTPUT_DIR       = PROJECT_ROOT / "results" / DATASET_NAME
COMBINED_H5      = OUTPUT_DIR / f"combined_{DATASET_NAME}.h5"


# -----------------------------
# SETTINGS
# -----------------------------
# PCA_COMPONENTS:  number of principal components to extract
# KMEANS_RUNS:     K-Means initializations per k for stability
# K_MIN / K_MAX:   range searched when auto-selecting optimal k
# K_INIT:          K-Means initializations used during k search
PCA_COMPONENTS = 6
KMEANS_RUNS    = 100
K_MIN          = 7
K_MAX          = 8
K_INIT         = 50


# -----------------------------
# HELPERS
# -----------------------------
def sample_id_from_filename(h5_name: str) -> str:
    """
    Extract the sample ID prefix from an H5 filename.
    CSO_007_DuringGrowth_... .h5  ->  CSO_007
    VideosSTO N=-5_0_rs.h5        ->  STO N=-5_0
    """
    name = h5_name.replace(".h5", "")
    if name.startswith("CSO_"):
        parts = name.split("_")
        return parts[0] + "_" + parts[1]
    name = name.replace("Videos", "")
    name = re.sub(r"_rs$", "", name)
    return name


def load_manifest():
    """
    Load target labels from the manifest file (CSV or Excel).
    For Excel files, reads only highlighted rows (green fill = FF00FF00).
    Returns a dict mapping sample_id -> target value (float).
    """
    path = MANIFEST_FILE
    if path.suffix in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        id_col  = headers.index("Sample #")
        tgt_col = headers.index(TARGET_COLUMN)
        result  = {}
        for row in ws.iter_rows(min_row=2):
            fill = row[tgt_col].fill.fgColor.rgb if row[tgt_col].fill else "00000000"
            if fill != "FF00FF00":
                continue
            sid = str(row[id_col].value).strip()
            val = row[tgt_col].value
            if val is not None:
                try:
                    result[sid] = float(val)
                except (ValueError, TypeError):
                    pass
        return result
    else:
        df = pd.read_csv(path)
        df["Sample_Name"] = df["Sample_Name"].astype(str).str.strip()
        return dict(zip(df["Sample_Name"], df[TARGET_COLUMN]))


def find_best_k(combined_h5_path, sample_ids, start_frames, end_frames, target_values):
    """
    Run K-Means for K_MIN to K_MAX, pick the k with the highest absolute
    Pearson correlation between the best cluster fraction and the target values.
    Returns the best k as an integer.
    """
    from sklearn.cluster import KMeans
    from scipy.stats import pearsonr

    with h5.File(combined_h5_path, "r") as f:
        eigenvalues = f["pca/eigenvalues"][:]

    best_k    = K_MIN
    best_corr = 0.0

    for k in range(K_MIN, K_MAX + 1):
        km = KMeans(n_clusters=k, n_init=K_INIT, max_iter=500, random_state=42)
        labels = km.fit_predict(eigenvalues)
        n_videos = len(start_frames)
        composition = np.zeros((n_videos, k))
        for i in range(n_videos):
            seg = labels[int(start_frames[i]):int(end_frames[i]) + 1]
            for c in range(k):
                composition[i, c] = np.mean(seg == c)
        max_abs = 0.0
        for c in range(k):
            col = composition[:, c]
            if col.std() < 1e-10:
                continue
            r, _ = pearsonr(col, target_values)
            if abs(r) > max_abs:
                max_abs = abs(r)
        print(f"  k={k:2d}  best |r| = {max_abs:.4f}")
        if max_abs > best_corr:
            best_corr = max_abs
            best_k    = k

    print(f"\nAuto-selected k = {best_k}  (|r| = {best_corr:.4f})")
    return best_k


# -----------------------------
# MAIN
# -----------------------------
def main():
    """
    Run the full combine + PCA + K-Means pipeline:
        1. Load frames from all individual H5 files, skip any without a
           matching label in the manifest.
        2. Concatenate all frames, recording per-video frame ranges and targets.
        3. Write combined_{DATASET_NAME}.h5 with a frame_map group.
        4. Run PCA on the combined frame stack.
        5. Auto-select the best K by target correlation.
        6. Run K-Means with the selected K.
    All results are written into combined_{DATASET_NAME}.h5 under results/DATASET_NAME/.
    """
    if not PREPROCESSED_DIR.exists():
        raise FileNotFoundError(f"Not found: {PREPROCESSED_DIR}")

    h5_files = sorted(PREPROCESSED_DIR.glob("*.h5"))
    if not h5_files:
        raise FileNotFoundError(f"No H5 files in {PREPROCESSED_DIR}")

    print(f"Dataset:        {DATASET_NAME}")
    print(f"Target column:  {TARGET_COLUMN}")
    print(f"Found {len(h5_files)} preprocessed H5 files.")

    target_lookup = load_manifest()
    print(f"Loaded {len(target_lookup)} labels from manifest.")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # -----------------------------
    # STEP 1: load all frames, build mapping
    # -----------------------------
    print("\n--- Loading frames from all videos ---")

    all_frames   = []
    all_times    = []
    sample_ids   = []
    start_frames = []
    end_frames   = []
    target_values = []

    total_frames_so_far = 0
    first_attrs         = None

    for i, h5_path in enumerate(h5_files, start=1):
        sample_id = sample_id_from_filename(h5_path.name)

        if sample_id not in target_lookup:
            print(f"[{i:2d}] {sample_id}: NO LABEL, skipping")
            continue

        with h5.File(h5_path, "r") as f:
            dset   = f["data/image_data"]
            frames = dset[:]
            times  = f["data/times"][:]
            attrs  = dict(dset.attrs)

        if first_attrs is None:
            first_attrs    = attrs
            expected_shape = frames.shape[:2]
            print(f"Expected frame shape: {expected_shape}")

        if frames.shape[:2] != expected_shape:
            print(f"[{i:2d}] {sample_id}: shape mismatch {frames.shape[:2]}, SKIPPING")
            continue

        n = frames.shape[2]
        all_frames.append(frames)
        all_times.append(times)
        sample_ids.append(sample_id)
        start_frames.append(total_frames_so_far)
        end_frames.append(total_frames_so_far + n - 1)
        target_values.append(float(target_lookup[sample_id]))
        total_frames_so_far += n

        print(f"[{i:2d}] {sample_id}: {n} frames, target={target_lookup[sample_id]:.4f}")

    # -----------------------------
    # STEP 2: concatenate
    # -----------------------------
    print(f"\nConcatenating {total_frames_so_far} total frames...")
    combined_frames = np.concatenate(all_frames, axis=2)
    combined_times  = np.concatenate(all_times)
    print(f"Combined shape: {combined_frames.shape}")

    combined_frames = combined_frames.astype(np.uint8)
    combined_frames[combined_frames == 0] = 1

    # -----------------------------
    # STEP 3: write combined H5
    # -----------------------------
    print(f"\nWriting combined H5: {COMBINED_H5}")
    with h5.File(COMBINED_H5, "w") as f:
        data_grp = f.create_group("data")
        dset     = data_grp.create_dataset("image_data", data=combined_frames)
        data_grp.create_dataset("times", data=combined_times)

        for k, v in first_attrs.items():
            dset.attrs[k] = v
        dset.attrs["total_frames"] = total_frames_so_far
        dset.attrs["x_res"]        = combined_frames.shape[1]
        dset.attrs["y_res"]        = combined_frames.shape[0]

        map_grp = data_grp.create_group("frame_map")
        map_grp.create_dataset("sample_ids",    data=np.array(sample_ids, dtype="S"))
        map_grp.create_dataset("start_frames",  data=np.array(start_frames,  dtype=np.int64))
        map_grp.create_dataset("end_frames",    data=np.array(end_frames,    dtype=np.int64))
        map_grp.create_dataset("xps_values",    data=np.array(target_values, dtype=np.float64))
        map_grp.create_dataset("target_values", data=np.array(target_values, dtype=np.float64))

    print("Combined H5 written.")

    # -----------------------------
    # STEP 4: run PCA
    # -----------------------------
    print("\n--- Running PCA ---")
    pca(str(COMBINED_H5), PCA_COMPONENTS)

    # -----------------------------
    # STEP 5: auto-select optimal K
    # -----------------------------
    print("\n--- Finding optimal K ---")
    best_k = find_best_k(
        str(COMBINED_H5), sample_ids,
        start_frames, end_frames, target_values
    )

    # -----------------------------
    # STEP 6: run K-Means with best K
    # -----------------------------
    print(f"\n--- Running K-Means with k={best_k} ---")
    k_means(str(COMBINED_H5), KMEANS_RUNS, (best_k, best_k))

    print("\n" + "=" * 60)
    print("DONE")
    print("=" * 60)
    print(f"Dataset:         {DATASET_NAME}")
    print(f"Output:          {COMBINED_H5}")
    print(f"Videos combined: {len(sample_ids)}")
    print(f"Total frames:    {total_frames_so_far}")
    print(f"PCA components:  {PCA_COMPONENTS}")
    print(f"K-Means k:       {best_k}  (auto-selected)")


if __name__ == "__main__":
    main()