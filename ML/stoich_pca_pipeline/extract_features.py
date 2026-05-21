"""
Script 4: extract per-video features from the combined H5 for regression.

Reads combined_{DATASET_NAME}.h5, slices PCA eigenvalues and K-Means cluster
labels by each video's frame range, computes summary features, and writes
a CSV ready for a regressor.

Edit DATASET_NAME below to switch datasets.
For CSO, lattice parameters from the manifest are included as extra features.

Features per video:
    - pc{1..6}_mean        : mean of each PCA component across the video
    - pc{1..6}_std         : std of each PCA component across the video
    - pc{1..6}_final       : final eigenvalue (last frame) per component
    - cluster_frac_{1..k}  : fraction of frames in each K-Means cluster
    - dominant_cluster     : the cluster with the most frames
    - lattice_param        : out-of-plane lattice parameter (CSO only)
    - target_value         : target to predict (2-theta for CSO, XPS for STO)
"""

from pathlib import Path
import h5py as h5
import numpy as np
import pandas as pd


PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent


# -----------------------------
# DATASET SELECTION — edit this to switch datasets
# -----------------------------
DATASET_NAME = "CSO"   # "STO" or "CSO"


# -----------------------------
# PATHS — derived automatically from DATASET_NAME
# -----------------------------
COMBINED_H5  = PROJECT_ROOT / "results" / DATASET_NAME / f"combined_{DATASET_NAME}.h5"
OUTPUT_CSV   = PROJECT_ROOT / "results" / DATASET_NAME / "per_video_features.csv"
MANIFEST_FILE = PROJECT_ROOT / "stoich_dataset" / DATASET_NAME / "labels" / "CSO log.xlsx"

# Columns to read from manifest as extra features (set to [] for STO)
EXTRA_FEATURE_COLUMNS = []   # lattice parameter is now the target, not a feature

N_COMPONENTS = 6   # must match PCA_COMPONENTS in combine_and_pca.py


def load_extra_features():
    """
    Load extra per-sample features from the manifest file.
    For CSO: reads lattice parameter from highlighted rows only.
    Returns a dict mapping sample_id -> dict of feature_name -> float.
    """
    if not EXTRA_FEATURE_COLUMNS:
        return {}

    path = MANIFEST_FILE
    if path.suffix in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(str(path))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        id_col   = headers.index("Sample #")
        feat_cols = [headers.index(col) for col in EXTRA_FEATURE_COLUMNS]
        result = {}
        for row in ws.iter_rows(min_row=2):
            # only highlighted rows
            fill = row[feat_cols[0]].fill.fgColor.rgb if row[feat_cols[0]].fill else "00000000"
            if fill != "FF00FF00":
                continue
            sid = str(row[id_col].value).strip()
            feats = {}
            for col_idx, col_name in zip(feat_cols, EXTRA_FEATURE_COLUMNS):
                val = row[col_idx].value
                try:
                    feats[col_name] = float(val)
                except (ValueError, TypeError):
                    feats[col_name] = float("nan")
            result[sid] = feats
        return result
    else:
        df = pd.read_csv(path)
        result = {}
        for _, row in df.iterrows():
            sid = str(row["Sample_Name"]).strip()
            result[sid] = {col: float(row[col]) for col in EXTRA_FEATURE_COLUMNS}
        return result


def main():
    """
    Extract per-video summary features from the combined H5 and save to CSV.
    Reads K from the H5 file automatically so it works for any k value.
    Lattice parameters (CSO) or other extra features from the manifest
    are appended as additional columns.
    """
    if not COMBINED_H5.exists():
        raise FileNotFoundError(f"Not found: {COMBINED_H5}")

    print(f"Dataset:  {DATASET_NAME}")
    print(f"Reading:  {COMBINED_H5}\n")

    with h5.File(COMBINED_H5, "r") as f:
        eigenvalues = f["pca/eigenvalues"][:]

        # Read K automatically from whatever k was saved
        kmeans_keys = list(f["kmeans"].keys())
        k_key = kmeans_keys[-1]   # e.g. "k=11"
        N_CLUSTERS = int(k_key.split("=")[1])
        print(f"Using K-Means key: {k_key}  (k={N_CLUSTERS})")

        kmeans_labels = f[f"kmeans/{k_key}/labels"][:]

        fmap         = f["data/frame_map"]
        sample_ids   = [s.decode() if isinstance(s, bytes) else s
                        for s in fmap["sample_ids"][:]]
        start_frames = fmap["start_frames"][:]
        end_frames   = fmap["end_frames"][:]
        target_values = fmap["target_values"][:]

    print(f"Total frames:     {eigenvalues.shape[0]}")
    print(f"PCA components:   {eigenvalues.shape[1]}")
    print(f"Videos:           {len(sample_ids)}\n")

    extra_features = load_extra_features()
    if extra_features:
        print(f"Extra features loaded for {len(extra_features)} samples: {EXTRA_FEATURE_COLUMNS}\n")

    rows = []

    for i, sid in enumerate(sample_ids):
        start = int(start_frames[i])
        end   = int(end_frames[i])

        video_eigs   = eigenvalues[start:end + 1]
        video_labels = kmeans_labels[start:end + 1]

        pc_means  = video_eigs.mean(axis=0)
        pc_stds   = video_eigs.std(axis=0)
        pc_finals = video_eigs[-1]

        cluster_fracs = np.zeros(N_CLUSTERS, dtype=np.float64)
        for k in range(N_CLUSTERS):
            cluster_fracs[k] = np.mean(video_labels == k)
        dominant = int(np.argmax(cluster_fracs))

        row = {"sample_id": sid}
        for k in range(N_COMPONENTS):
            row[f"pc{k+1}_mean"]  = float(pc_means[k])
            row[f"pc{k+1}_std"]   = float(pc_stds[k])
            row[f"pc{k+1}_final"] = float(pc_finals[k])
        for k in range(N_CLUSTERS):
            row[f"cluster_frac_{k+1}"] = float(cluster_fracs[k])
        row["dominant_cluster"] = dominant

        # Add extra features from manifest (e.g. lattice parameter for CSO)
        if sid in extra_features:
            for feat_name, feat_val in extra_features[sid].items():
                row[feat_name] = feat_val

        row["target_value"] = float(target_values[i])

        rows.append(row)
        print(f"[{i+1:2d}] {sid}: {end - start + 1} frames, "
              f"dominant_cluster={dominant+1}, target={target_values[i]:.4f}")

    df = pd.DataFrame(rows)
    df.to_csv(OUTPUT_CSV, index=False)

    print(f"\nSaved: {OUTPUT_CSV}")
    print(f"Shape: {df.shape}")
    print(f"Columns: {list(df.columns)}")
    print(f"Target range: {df['target_value'].min():.4f} - {df['target_value'].max():.4f}")


if __name__ == "__main__":
    main()